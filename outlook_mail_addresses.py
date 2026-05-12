from __future__ import annotations

import csv
import json
import shutil
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterator

from outlook_contacts import (
    PROJECT_ROOT,
    _normalize,
    _safe_iso,
    _safe_property,
    _safe_text,
    _truncate,
    outlook_session,
)


DEFAULT_MAIL_OUTPUT_DIR = PROJECT_ROOT / "output" / "mail-addresses" / "latest"
DEFAULT_MAIL_DAYS_BACK = 365

OL_FOLDER_INBOX = 6
OL_FOLDER_SENT_MAIL = 5
OL_CLASS_MAIL = 43

PR_SMTP_ADDRESS = "http://schemas.microsoft.com/mapi/proptag/0x39FE001E"


MAIL_ADDRESS_COLUMNS = [
    "email_address",
    "primary_display_name",
    "display_names",
    "incoming_sender_count",
    "incoming_recipient_count",
    "outgoing_sender_count",
    "outgoing_recipient_count",
    "message_count_total",
    "first_seen_at",
    "last_seen_at",
    "stores",
    "folders",
    "sample_subjects",
]


@dataclass(slots=True)
class MailAddressFilters:
    store_name: str | None = None
    folder_path: str | None = None
    include_subfolders: bool = True
    scan_inbox: bool = True
    scan_sent_mail: bool = True
    days_back: int | None = DEFAULT_MAIL_DAYS_BACK
    max_messages: int | None = None
    address_scope: str = "correspondents"


@dataclass(slots=True)
class AddressAggregate:
    email_address: str
    display_name_counts: Counter[str] = field(default_factory=Counter)
    incoming_sender_count: int = 0
    incoming_recipient_count: int = 0
    outgoing_sender_count: int = 0
    outgoing_recipient_count: int = 0
    first_seen_at: datetime | None = None
    last_seen_at: datetime | None = None
    stores: set[str] = field(default_factory=set)
    folders: set[str] = field(default_factory=set)
    sample_subjects: list[str] = field(default_factory=list)

    def record(
        self,
        *,
        display_name: str,
        role: str,
        seen_at: datetime | None,
        store_name: str,
        folder_path: str,
        subject: str,
    ) -> None:
        if display_name:
            self.display_name_counts[display_name] += 1
        if role == "incoming_sender":
            self.incoming_sender_count += 1
        elif role == "incoming_recipient":
            self.incoming_recipient_count += 1
        elif role == "outgoing_sender":
            self.outgoing_sender_count += 1
        elif role == "outgoing_recipient":
            self.outgoing_recipient_count += 1

        if seen_at is not None:
            if self.first_seen_at is None or seen_at < self.first_seen_at:
                self.first_seen_at = seen_at
            if self.last_seen_at is None or seen_at > self.last_seen_at:
                self.last_seen_at = seen_at

        if store_name:
            self.stores.add(store_name)
        if folder_path:
            self.folders.add(folder_path)
        subject = subject.strip()
        if subject and subject not in self.sample_subjects and len(self.sample_subjects) < 5:
            self.sample_subjects.append(subject)

    def as_row(self) -> dict[str, str]:
        primary_display_name = ""
        display_names = sorted(
            self.display_name_counts,
            key=lambda name: (-self.display_name_counts[name], name.lower()),
        )
        if display_names:
            primary_display_name = display_names[0]

        message_count_total = (
            self.incoming_sender_count
            + self.incoming_recipient_count
            + self.outgoing_sender_count
            + self.outgoing_recipient_count
        )

        return {
            "email_address": self.email_address,
            "primary_display_name": primary_display_name,
            "display_names": "; ".join(display_names),
            "incoming_sender_count": str(self.incoming_sender_count),
            "incoming_recipient_count": str(self.incoming_recipient_count),
            "outgoing_sender_count": str(self.outgoing_sender_count),
            "outgoing_recipient_count": str(self.outgoing_recipient_count),
            "message_count_total": str(message_count_total),
            "first_seen_at": _safe_iso(self.first_seen_at),
            "last_seen_at": _safe_iso(self.last_seen_at),
            "stores": "; ".join(sorted(self.stores, key=str.lower)),
            "folders": "; ".join(sorted(self.folders, key=str.lower)),
            "sample_subjects": "; ".join(self.sample_subjects),
        }


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize_email(value: str | None) -> str:
    return (value or "").strip().lower()


def _is_probable_email(value: str | None) -> bool:
    text = (value or "").strip()
    return "@" in text and " " not in text


def _property_accessor_value(obj: Any, property_name: str) -> str:
    accessor = _safe_property(obj, "PropertyAccessor")
    if accessor is None:
        return ""
    try:
        return _safe_text(accessor.GetProperty(property_name))
    except Exception:
        return ""


def _resolve_address_entry_smtp(address_entry: Any) -> str:
    if address_entry is None:
        return ""

    smtp = _property_accessor_value(address_entry, PR_SMTP_ADDRESS)
    if _is_probable_email(smtp):
        return smtp

    try:
        exchange_user = address_entry.GetExchangeUser()
    except Exception:
        exchange_user = None
    smtp = _safe_text(_safe_property(exchange_user, "PrimarySmtpAddress"))
    if _is_probable_email(smtp):
        return smtp

    try:
        exchange_distribution_list = address_entry.GetExchangeDistributionList()
    except Exception:
        exchange_distribution_list = None
    smtp = _safe_text(_safe_property(exchange_distribution_list, "PrimarySmtpAddress"))
    if _is_probable_email(smtp):
        return smtp

    raw_address = _safe_text(_safe_property(address_entry, "Address"))
    if _is_probable_email(raw_address):
        return raw_address
    return ""


def _resolve_recipient_address(recipient: Any) -> tuple[str, str]:
    display_name = _safe_text(_safe_property(recipient, "Name"))
    smtp = _property_accessor_value(recipient, PR_SMTP_ADDRESS)
    if not _is_probable_email(smtp):
        smtp = _resolve_address_entry_smtp(_safe_property(recipient, "AddressEntry"))
    if not _is_probable_email(smtp):
        raw_address = _safe_text(_safe_property(recipient, "Address"))
        smtp = raw_address if _is_probable_email(raw_address) else ""
    return display_name, smtp


def _resolve_sender_address(mail_item: Any) -> tuple[str, str]:
    display_name = _safe_text(_safe_property(mail_item, "SenderName"))
    raw_address = _safe_text(_safe_property(mail_item, "SenderEmailAddress"))
    sender_email_type = _safe_text(_safe_property(mail_item, "SenderEmailType")).upper()
    if sender_email_type != "EX" and _is_probable_email(raw_address):
        return display_name, raw_address

    smtp = _resolve_address_entry_smtp(_safe_property(mail_item, "Sender"))
    if _is_probable_email(smtp):
        return display_name, smtp
    if _is_probable_email(raw_address):
        return display_name, raw_address
    return display_name, ""


def _message_datetime(mail_item: Any, direction: str) -> datetime | None:
    if direction == "incoming":
        value = _safe_property(mail_item, "ReceivedTime")
    else:
        value = _safe_property(mail_item, "SentOn")
    return value if isinstance(value, datetime) else None


def _passes_days_back(seen_at: datetime | None, days_back: int | None) -> bool:
    if seen_at is None or days_back is None:
        return True
    if seen_at.tzinfo is None:
        cutoff = datetime.now() - timedelta(days=days_back)
    else:
        cutoff = datetime.now(seen_at.tzinfo) - timedelta(days=days_back)
    return seen_at >= cutoff


def _iter_mail_root_folders(
    session: Any,
    *,
    store_name: str | None,
    scan_inbox: bool,
    scan_sent_mail: bool,
) -> Iterator[tuple[dict[str, Any], Any]]:
    stores = getattr(session, "Stores", None)
    if stores is None:
        return

    root_specs: list[tuple[int, str, str]] = []
    if scan_inbox:
        root_specs.append((OL_FOLDER_INBOX, "inbox", "incoming"))
    if scan_sent_mail:
        root_specs.append((OL_FOLDER_SENT_MAIL, "sent_mail", "outgoing"))

    store_count = int(getattr(stores, "Count", 0) or 0)
    for store_index in range(1, store_count + 1):
        store = stores.Item(store_index)
        current_store_name = _safe_text(_safe_property(store, "DisplayName")) or f"Store {store_index}"
        if store_name and _normalize(current_store_name) != _normalize(store_name):
            continue

        for folder_type, root_kind, direction in root_specs:
            try:
                root_folder = store.GetDefaultFolder(folder_type)
            except Exception:
                continue
            if root_folder is None:
                continue
            root_name = _safe_text(_safe_property(root_folder, "Name")) or root_kind
            root_meta = {
                "store_name": current_store_name,
                "folder_name": root_name,
                "folder_path": root_name,
                "root_kind": root_kind,
                "direction": direction,
                "entry_id": _safe_text(_safe_property(root_folder, "EntryID")),
            }
            yield root_meta, root_folder


def _walk_mail_folders(
    folder: Any,
    folder_meta: dict[str, Any],
    *,
    include_subfolders: bool,
) -> Iterator[tuple[dict[str, Any], Any]]:
    yield folder_meta, folder
    if not include_subfolders:
        return

    subfolders = _safe_property(folder, "Folders")
    subfolder_count = int(getattr(subfolders, "Count", 0) or 0)
    for index in range(1, subfolder_count + 1):
        child = subfolders.Item(index)
        child_name = _safe_text(_safe_property(child, "Name")) or f"Folder {index}"
        child_meta = {
            **folder_meta,
            "folder_name": child_name,
            "folder_path": f"{folder_meta['folder_path']}/{child_name}",
            "entry_id": _safe_text(_safe_property(child, "EntryID")),
        }
        yield from _walk_mail_folders(child, child_meta, include_subfolders=include_subfolders)


def collect_mail_folders(filters: MailAddressFilters | None = None) -> list[dict[str, Any]]:
    filters = filters or MailAddressFilters()
    folders: list[dict[str, Any]] = []

    with outlook_session() as session:
        for root_meta, root_folder in _iter_mail_root_folders(
            session,
            store_name=filters.store_name,
            scan_inbox=filters.scan_inbox,
            scan_sent_mail=filters.scan_sent_mail,
        ):
            for folder_meta, folder in _walk_mail_folders(
                root_folder,
                root_meta,
                include_subfolders=filters.include_subfolders,
            ):
                if filters.folder_path and _normalize(folder_meta["folder_path"]) != _normalize(filters.folder_path):
                    continue
                items = _safe_property(folder, "Items")
                folders.append(
                    {
                        **folder_meta,
                        "item_count": int(getattr(items, "Count", 0) or 0),
                    }
                )

    folders.sort(
        key=lambda item: (
            item["store_name"].lower(),
            item["direction"].lower(),
            item["folder_path"].lower(),
        )
    )
    return folders


def _record_address(
    aggregates: dict[str, AddressAggregate],
    *,
    email_address: str,
    display_name: str,
    role: str,
    seen_at: datetime | None,
    store_name: str,
    folder_path: str,
    subject: str,
) -> None:
    normalized_email = _normalize_email(email_address)
    if not _is_probable_email(normalized_email):
        return
    aggregate = aggregates.get(normalized_email)
    if aggregate is None:
        aggregate = AddressAggregate(email_address=normalized_email)
        aggregates[normalized_email] = aggregate
    aggregate.record(
        display_name=display_name,
        role=role,
        seen_at=seen_at,
        store_name=store_name,
        folder_path=folder_path,
        subject=subject,
    )


def _record_mail_participants(
    aggregates: dict[str, AddressAggregate],
    mail_item: Any,
    *,
    direction: str,
    scope: str,
    store_name: str,
    folder_path: str,
) -> None:
    seen_at = _message_datetime(mail_item, direction)
    subject = _safe_text(_safe_property(mail_item, "Subject"))

    include_sender = scope == "all-participants" or direction == "incoming"
    include_recipients = scope == "all-participants" or direction == "outgoing"

    if include_sender:
        display_name, email_address = _resolve_sender_address(mail_item)
        _record_address(
            aggregates,
            email_address=email_address,
            display_name=display_name,
            role=f"{direction}_sender",
            seen_at=seen_at,
            store_name=store_name,
            folder_path=folder_path,
            subject=subject,
        )

    if include_recipients:
        recipients = _safe_property(mail_item, "Recipients")
        recipient_count = int(getattr(recipients, "Count", 0) or 0)
        for index in range(1, recipient_count + 1):
            try:
                recipient = recipients.Item(index)
            except Exception:
                continue
            display_name, email_address = _resolve_recipient_address(recipient)
            _record_address(
                aggregates,
                email_address=email_address,
                display_name=display_name,
                role=f"{direction}_recipient",
                seen_at=seen_at,
                store_name=store_name,
                folder_path=folder_path,
                subject=subject,
            )


def collect_mail_address_rows(filters: MailAddressFilters | None = None) -> tuple[list[dict[str, str]], dict[str, Any]]:
    filters = filters or MailAddressFilters()
    aggregates: dict[str, AddressAggregate] = {}
    scanned_message_count = 0

    with outlook_session() as session:
        stop_scanning = False
        for root_meta, root_folder in _iter_mail_root_folders(
            session,
            store_name=filters.store_name,
            scan_inbox=filters.scan_inbox,
            scan_sent_mail=filters.scan_sent_mail,
        ):
            if stop_scanning:
                break
            for folder_meta, folder in _walk_mail_folders(
                root_folder,
                root_meta,
                include_subfolders=filters.include_subfolders,
            ):
                if stop_scanning:
                    break
                if filters.folder_path and _normalize(folder_meta["folder_path"]) != _normalize(filters.folder_path):
                    continue

                items = _safe_property(folder, "Items")
                item_count = int(getattr(items, "Count", 0) or 0)
                for index in range(1, item_count + 1):
                    if filters.max_messages is not None and scanned_message_count >= filters.max_messages:
                        stop_scanning = True
                        break
                    try:
                        item = items.Item(index)
                    except Exception:
                        continue
                    if _safe_property(item, "Class") != OL_CLASS_MAIL:
                        continue

                    seen_at = _message_datetime(item, folder_meta["direction"])
                    if not _passes_days_back(seen_at, filters.days_back):
                        continue

                    scanned_message_count += 1
                    _record_mail_participants(
                        aggregates,
                        item,
                        direction=folder_meta["direction"],
                        scope=filters.address_scope,
                        store_name=folder_meta["store_name"],
                        folder_path=folder_meta["folder_path"],
                    )

    rows = [aggregate.as_row() for aggregate in aggregates.values()]
    rows.sort(
        key=lambda row: (
            row["primary_display_name"].lower(),
            row["email_address"].lower(),
        )
    )
    stats = {
        "scanned_message_count": scanned_message_count,
        "unique_address_count": len(rows),
        "address_scope": filters.address_scope,
    }
    return rows, stats


def preview_mail_addresses(
    *,
    filters: MailAddressFilters | None = None,
    limit: int = 25,
    offset: int = 0,
    include_subject_preview: bool = True,
    preview_chars: int = 200,
) -> dict[str, Any]:
    rows, stats = collect_mail_address_rows(filters)
    page = rows[offset : offset + max(1, min(limit, 100))]
    preview_rows: list[dict[str, str]] = []
    for row in page:
        item = dict(row)
        if include_subject_preview:
            item["sample_subjects_preview"] = _truncate(item.pop("sample_subjects", ""), preview_chars)
        else:
            item.pop("sample_subjects", None)
        preview_rows.append(item)
    return {
        "count": len(preview_rows),
        "total": len(rows),
        "offset": offset,
        "limit": limit,
        "scanned_message_count": stats["scanned_message_count"],
        "addresses": preview_rows,
    }


def search_mail_address_rows(
    query: str,
    filters: MailAddressFilters | None = None,
    *,
    limit: int = 25,
) -> dict[str, Any]:
    normalized_query = _normalize(query)
    if not normalized_query:
        return {"count": 0, "results": [], "scanned_message_count": 0}

    rows, stats = collect_mail_address_rows(filters)
    results: list[dict[str, str]] = []
    for row in rows:
        haystack = "\n".join(
            [
                row["email_address"],
                row["primary_display_name"],
                row["display_names"],
                row["stores"],
                row["folders"],
                row["sample_subjects"],
            ]
        ).lower()
        if normalized_query in haystack:
            results.append(row)
            if len(results) >= max(1, min(limit, 100)):
                break

    return {
        "count": len(results),
        "results": results,
        "scanned_message_count": stats["scanned_message_count"],
    }


def get_mail_address(
    *,
    email_address: str,
    filters: MailAddressFilters | None = None,
) -> dict[str, Any]:
    normalized_email = _normalize_email(email_address)
    rows, stats = collect_mail_address_rows(filters)
    matches = [row for row in rows if row["email_address"] == normalized_email]
    return {
        "match_count": len(matches),
        "matches": matches[:20],
        "scanned_message_count": stats["scanned_message_count"],
    }


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=MAIL_ADDRESS_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Mail Addresses"
    sheet.append(MAIL_ADDRESS_COLUMNS)

    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for column_index, header in enumerate(MAIL_ADDRESS_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = header_font

    for row in rows:
        sheet.append([row.get(column, "") for column in MAIL_ADDRESS_COLUMNS])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_index, header in enumerate(MAIL_ADDRESS_COLUMNS, start=1):
        column_letter = get_column_letter(column_index)
        max_length = len(header)
        for cell in sheet[column_letter]:
            cell.alignment = wrap_alignment
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    workbook.save(path)


def _write_summary(
    path: Path,
    rows: list[dict[str, str]],
    filters: MailAddressFilters,
    stats: dict[str, Any],
) -> None:
    store_counts = Counter()
    folder_counts = Counter()
    for row in rows:
        for store_name in [item.strip() for item in row["stores"].split(";") if item.strip()]:
            store_counts[store_name] += 1
        for folder_name in [item.strip() for item in row["folders"].split(";") if item.strip()]:
            folder_counts[folder_name] += 1

    lines = [
        "# Outlook Mail Address Export Summary",
        "",
        f"- Generated at: `{utc_now_iso()}`",
        f"- Unique addresses: `{len(rows)}`",
        f"- Messages scanned: `{stats['scanned_message_count']}`",
        f"- Address scope: `{filters.address_scope}`",
        f"- Store filter: `{filters.store_name or 'all'}`",
        f"- Folder filter: `{filters.folder_path or 'all'}`",
        f"- Include subfolders: `{'yes' if filters.include_subfolders else 'no'}`",
        f"- Scan inbox: `{'yes' if filters.scan_inbox else 'no'}`",
        f"- Scan sent mail: `{'yes' if filters.scan_sent_mail else 'no'}`",
        f"- Days back: `{filters.days_back if filters.days_back is not None else 'all'}`",
        f"- Max messages: `{filters.max_messages if filters.max_messages is not None else 'unlimited'}`",
        "",
        "## Stores",
        "",
    ]

    if store_counts:
        for store_name, count in sorted(store_counts.items(), key=lambda item: item[0].lower()):
            lines.append(f"- `{store_name}`: `{count}` addresses")
    else:
        lines.append("- No addresses exported.")

    lines.extend(["", "## Folders", ""])
    if folder_counts:
        for folder_name, count in sorted(folder_counts.items(), key=lambda item: item[0].lower()):
            lines.append(f"- `{folder_name}`: `{count}` addresses")
    else:
        lines.append("- No folders exported.")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def export_mail_addresses_snapshot(
    *,
    output_dir: Path = DEFAULT_MAIL_OUTPUT_DIR,
    filters: MailAddressFilters | None = None,
    clean: bool = False,
) -> dict[str, Any]:
    filters = filters or MailAddressFilters()

    if clean and output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    rows, stats = collect_mail_address_rows(filters)

    json_path = output_dir / "mail_addresses.json"
    csv_path = output_dir / "mail_addresses.csv"
    xlsx_path = output_dir / "mail_addresses.xlsx"
    manifest_path = output_dir / "manifest.json"
    summary_path = output_dir / "summary.md"

    _write_json(json_path, rows)
    _write_csv(csv_path, rows)
    _write_xlsx(xlsx_path, rows)
    _write_summary(summary_path, rows, filters, stats)

    manifest = {
        "generated_at": utc_now_iso(),
        "unique_address_count": len(rows),
        "scanned_message_count": stats["scanned_message_count"],
        "store_filter": filters.store_name,
        "folder_filter": filters.folder_path,
        "include_subfolders": filters.include_subfolders,
        "scan_inbox": filters.scan_inbox,
        "scan_sent_mail": filters.scan_sent_mail,
        "days_back": filters.days_back,
        "max_messages": filters.max_messages,
        "address_scope": filters.address_scope,
        "files": {
            "mail_addresses_json": str(json_path),
            "mail_addresses_csv": str(csv_path),
            "mail_addresses_xlsx": str(xlsx_path),
            "summary_md": str(summary_path),
        },
    }
    _write_json(manifest_path, manifest)
    manifest["files"]["manifest_json"] = str(manifest_path)
    return manifest
