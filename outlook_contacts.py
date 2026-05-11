from __future__ import annotations

import csv
import json
import platform
import shutil
from collections import Counter
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator


PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output" / "latest"

OL_FOLDER_CONTACTS = 10
OL_CLASS_CONTACT = 40
OL_CLASS_DISTLIST = 69


CONTACT_COLUMNS = [
    "entry_id",
    "item_type",
    "store_name",
    "folder_path",
    "created_at",
    "modified_at",
    "full_name",
    "file_as",
    "title",
    "company_name",
    "job_title",
    "department",
    "assistant_name",
    "manager_name",
    "nickname",
    "spouse",
    "children",
    "birthday",
    "anniversary",
    "email_1_display_name",
    "email_1_address",
    "email_1_type",
    "email_2_display_name",
    "email_2_address",
    "email_2_type",
    "email_3_display_name",
    "email_3_address",
    "email_3_type",
    "business_phone",
    "business_phone_2",
    "company_main_phone",
    "assistant_phone",
    "mobile_phone",
    "primary_phone",
    "home_phone",
    "home_phone_2",
    "other_phone",
    "pager",
    "car_phone",
    "radio_phone",
    "business_street",
    "business_city",
    "business_state",
    "business_postal_code",
    "business_country",
    "home_street",
    "home_city",
    "home_state",
    "home_postal_code",
    "home_country",
    "other_street",
    "other_city",
    "other_state",
    "other_postal_code",
    "other_country",
    "categories",
    "web_page",
    "im_address",
    "notes",
    "member_count",
    "members",
]


class OutlookAccessError(RuntimeError):
    pass


@dataclass(slots=True)
class ContactFilters:
    store_name: str | None = None
    folder_path: str | None = None
    include_distribution_lists: bool = True


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_iso(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    try:
        return value.isoformat()
    except AttributeError:
        return _safe_text(value)


def _stringify_sequence(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, str):
        return value
    try:
        return "; ".join(str(item).strip() for item in value if str(item).strip())
    except TypeError:
        return _safe_text(value)


def _truncate(text: str, max_chars: int) -> str:
    if max_chars <= 0 or len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...[truncated]"


def _normalize(value: str | None) -> str:
    return (value or "").strip().lower()


def _require_windows() -> None:
    if platform.system() != "Windows":
        raise OutlookAccessError(
            "This project is meant to run on Windows 11 with classic Outlook for Windows."
        )


def _import_outlook_modules() -> tuple[Any, Any]:
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise OutlookAccessError(
            "pywin32 is not installed. Run scripts/bootstrap.ps1 on Windows first."
        ) from exc
    return pythoncom, win32com.client


@contextmanager
def outlook_session() -> Iterator[Any]:
    _require_windows()
    pythoncom, win32_client = _import_outlook_modules()
    pythoncom.CoInitialize()
    try:
        try:
            application = win32_client.Dispatch("Outlook.Application")
            session = application.Session
        except Exception as exc:
            raise OutlookAccessError(
                "Failed to open Outlook through COM. Install classic Outlook for Windows "
                "and sign into the profile you want to export."
            ) from exc
        yield session
    finally:
        pythoncom.CoUninitialize()


def collect_status() -> dict[str, Any]:
    status: dict[str, Any] = {
        "platform": platform.platform(),
        "is_windows": platform.system() == "Windows",
        "pywin32_available": False,
        "outlook_com_available": False,
        "store_count": 0,
        "guidance": (
            "Run this on Windows 11 with classic Outlook for Windows installed and "
            "signed into a profile."
        ),
    }
    if not status["is_windows"]:
        status["error"] = (
            "This machine is not Windows. Move the repo to the Windows 11 machine "
            "where classic Outlook is installed."
        )
        return status

    try:
        _import_outlook_modules()
    except OutlookAccessError as exc:
        status["error"] = str(exc)
        return status

    status["pywin32_available"] = True

    try:
        with outlook_session() as session:
            stores = getattr(session, "Stores", None)
            status["outlook_com_available"] = True
            status["store_count"] = int(getattr(stores, "Count", 0) or 0)
    except OutlookAccessError as exc:
        status["error"] = str(exc)

    return status


def _safe_property(item: Any, property_name: str) -> Any:
    try:
        return getattr(item, property_name)
    except Exception:
        return None


def _iter_contact_folders(session: Any, store_name: str | None = None) -> Iterator[tuple[dict[str, Any], Any]]:
    stores = getattr(session, "Stores", None)
    if stores is None:
        return

    store_count = int(getattr(stores, "Count", 0) or 0)
    for store_index in range(1, store_count + 1):
        store = stores.Item(store_index)
        current_store_name = _safe_text(_safe_property(store, "DisplayName")) or f"Store {store_index}"
        if store_name and _normalize(current_store_name) != _normalize(store_name):
            continue

        try:
            root_folder = store.GetDefaultFolder(OL_FOLDER_CONTACTS)
        except Exception:
            continue
        if root_folder is None:
            continue

        root_name = _safe_text(_safe_property(root_folder, "Name")) or "Contacts"
        yield from _walk_contact_folder(root_folder, current_store_name, root_name)


def _walk_contact_folder(folder: Any, store_name: str, folder_path: str) -> Iterator[tuple[dict[str, Any], Any]]:
    folder_meta = {
        "store_name": store_name,
        "folder_name": _safe_text(_safe_property(folder, "Name")) or folder_path.rsplit("/", 1)[-1],
        "folder_path": folder_path,
        "entry_id": _safe_text(_safe_property(folder, "EntryID")),
    }
    yield folder_meta, folder

    subfolders = _safe_property(folder, "Folders")
    subfolder_count = int(getattr(subfolders, "Count", 0) or 0)
    for index in range(1, subfolder_count + 1):
        child = subfolders.Item(index)
        child_name = _safe_text(_safe_property(child, "Name")) or f"Folder {index}"
        child_path = f"{folder_path}/{child_name}"
        yield from _walk_contact_folder(child, store_name, child_path)


def _iter_folder_items(folder: Any) -> Iterator[Any]:
    items = _safe_property(folder, "Items")
    item_count = int(getattr(items, "Count", 0) or 0)
    for index in range(1, item_count + 1):
        try:
            yield items.Item(index)
        except Exception:
            continue


def _matches_filters(folder_meta: dict[str, Any], filters: ContactFilters) -> bool:
    if filters.store_name and _normalize(folder_meta["store_name"]) != _normalize(filters.store_name):
        return False
    if filters.folder_path and _normalize(folder_meta["folder_path"]) != _normalize(filters.folder_path):
        return False
    return True


def _distribution_list_members(item: Any) -> tuple[int, str]:
    members: list[str] = []
    try:
        member_count = int(getattr(item, "MemberCount", 0) or 0)
    except Exception:
        return 0, ""

    for index in range(1, member_count + 1):
        try:
            member = item.GetMember(index)
        except Exception:
            continue
        name = _safe_text(_safe_property(member, "Name"))
        address = _safe_text(_safe_property(member, "Address"))
        if name and address:
            members.append(f"{name} <{address}>")
        elif name:
            members.append(name)
        elif address:
            members.append(address)

    return member_count, "; ".join(members)


def _base_row(folder_meta: dict[str, Any], item: Any, item_type: str) -> dict[str, str]:
    row = {column: "" for column in CONTACT_COLUMNS}
    row["entry_id"] = _safe_text(_safe_property(item, "EntryID"))
    row["item_type"] = item_type
    row["store_name"] = folder_meta["store_name"]
    row["folder_path"] = folder_meta["folder_path"]
    row["created_at"] = _safe_iso(_safe_property(item, "CreationTime"))
    row["modified_at"] = _safe_iso(_safe_property(item, "LastModificationTime"))
    return row


def _extract_contact_row(folder_meta: dict[str, Any], item: Any) -> dict[str, str]:
    row = _base_row(folder_meta, item, "contact")
    row["full_name"] = _safe_text(_safe_property(item, "FullName"))
    row["file_as"] = _safe_text(_safe_property(item, "FileAs"))
    row["title"] = _safe_text(_safe_property(item, "Title"))
    row["company_name"] = _safe_text(_safe_property(item, "CompanyName"))
    row["job_title"] = _safe_text(_safe_property(item, "JobTitle"))
    row["department"] = _safe_text(_safe_property(item, "Department"))
    row["assistant_name"] = _safe_text(_safe_property(item, "AssistantName"))
    row["manager_name"] = _safe_text(_safe_property(item, "ManagerName"))
    row["nickname"] = _safe_text(_safe_property(item, "NickName"))
    row["spouse"] = _safe_text(_safe_property(item, "Spouse"))
    row["children"] = _stringify_sequence(_safe_property(item, "Children"))
    row["birthday"] = _safe_iso(_safe_property(item, "Birthday"))
    row["anniversary"] = _safe_iso(_safe_property(item, "Anniversary"))
    row["email_1_display_name"] = _safe_text(_safe_property(item, "Email1DisplayName"))
    row["email_1_address"] = _safe_text(_safe_property(item, "Email1Address"))
    row["email_1_type"] = _safe_text(_safe_property(item, "Email1AddressType"))
    row["email_2_display_name"] = _safe_text(_safe_property(item, "Email2DisplayName"))
    row["email_2_address"] = _safe_text(_safe_property(item, "Email2Address"))
    row["email_2_type"] = _safe_text(_safe_property(item, "Email2AddressType"))
    row["email_3_display_name"] = _safe_text(_safe_property(item, "Email3DisplayName"))
    row["email_3_address"] = _safe_text(_safe_property(item, "Email3Address"))
    row["email_3_type"] = _safe_text(_safe_property(item, "Email3AddressType"))
    row["business_phone"] = _safe_text(_safe_property(item, "BusinessTelephoneNumber"))
    row["business_phone_2"] = _safe_text(_safe_property(item, "Business2TelephoneNumber"))
    row["company_main_phone"] = _safe_text(_safe_property(item, "CompanyMainTelephoneNumber"))
    row["assistant_phone"] = _safe_text(_safe_property(item, "AssistantTelephoneNumber"))
    row["mobile_phone"] = _safe_text(_safe_property(item, "MobileTelephoneNumber"))
    row["primary_phone"] = _safe_text(_safe_property(item, "PrimaryTelephoneNumber"))
    row["home_phone"] = _safe_text(_safe_property(item, "HomeTelephoneNumber"))
    row["home_phone_2"] = _safe_text(_safe_property(item, "Home2TelephoneNumber"))
    row["other_phone"] = _safe_text(_safe_property(item, "OtherTelephoneNumber"))
    row["pager"] = _safe_text(_safe_property(item, "PagerNumber"))
    row["car_phone"] = _safe_text(_safe_property(item, "CarTelephoneNumber"))
    row["radio_phone"] = _safe_text(_safe_property(item, "RadioTelephoneNumber"))
    row["business_street"] = _safe_text(_safe_property(item, "BusinessAddressStreet"))
    row["business_city"] = _safe_text(_safe_property(item, "BusinessAddressCity"))
    row["business_state"] = _safe_text(_safe_property(item, "BusinessAddressState"))
    row["business_postal_code"] = _safe_text(_safe_property(item, "BusinessAddressPostalCode"))
    row["business_country"] = _safe_text(_safe_property(item, "BusinessAddressCountry"))
    row["home_street"] = _safe_text(_safe_property(item, "HomeAddressStreet"))
    row["home_city"] = _safe_text(_safe_property(item, "HomeAddressCity"))
    row["home_state"] = _safe_text(_safe_property(item, "HomeAddressState"))
    row["home_postal_code"] = _safe_text(_safe_property(item, "HomeAddressPostalCode"))
    row["home_country"] = _safe_text(_safe_property(item, "HomeAddressCountry"))
    row["other_street"] = _safe_text(_safe_property(item, "OtherAddressStreet"))
    row["other_city"] = _safe_text(_safe_property(item, "OtherAddressCity"))
    row["other_state"] = _safe_text(_safe_property(item, "OtherAddressState"))
    row["other_postal_code"] = _safe_text(_safe_property(item, "OtherAddressPostalCode"))
    row["other_country"] = _safe_text(_safe_property(item, "OtherAddressCountry"))
    row["categories"] = _safe_text(_safe_property(item, "Categories"))
    row["web_page"] = _safe_text(_safe_property(item, "WebPage"))
    row["im_address"] = _safe_text(_safe_property(item, "IMAddress"))
    row["notes"] = _safe_text(_safe_property(item, "Body"))
    return row


def _extract_distribution_list_row(folder_meta: dict[str, Any], item: Any) -> dict[str, str]:
    row = _base_row(folder_meta, item, "distribution_list")
    row["full_name"] = _safe_text(_safe_property(item, "DLName")) or _safe_text(_safe_property(item, "Subject"))
    row["file_as"] = _safe_text(_safe_property(item, "FileAs"))
    row["categories"] = _safe_text(_safe_property(item, "Categories"))
    row["notes"] = _safe_text(_safe_property(item, "Body"))
    member_count, members = _distribution_list_members(item)
    row["member_count"] = str(member_count)
    row["members"] = members
    return row


def collect_contact_rows(filters: ContactFilters | None = None) -> list[dict[str, str]]:
    filters = filters or ContactFilters()
    rows: list[dict[str, str]] = []

    with outlook_session() as session:
        for folder_meta, folder in _iter_contact_folders(session, store_name=filters.store_name):
            if not _matches_filters(folder_meta, filters):
                continue

            for item in _iter_folder_items(folder):
                class_id = _safe_property(item, "Class")
                if class_id == OL_CLASS_CONTACT:
                    rows.append(_extract_contact_row(folder_meta, item))
                elif class_id == OL_CLASS_DISTLIST and filters.include_distribution_lists:
                    rows.append(_extract_distribution_list_row(folder_meta, item))

    rows.sort(
        key=lambda row: (
            row["full_name"].lower(),
            row["store_name"].lower(),
            row["folder_path"].lower(),
            row["entry_id"].lower(),
        )
    )
    return rows


def collect_contact_folders(store_name: str | None = None) -> list[dict[str, Any]]:
    folders: list[dict[str, Any]] = []
    with outlook_session() as session:
        for folder_meta, folder in _iter_contact_folders(session, store_name=store_name):
            contact_count = 0
            distribution_list_count = 0
            for item in _iter_folder_items(folder):
                class_id = _safe_property(item, "Class")
                if class_id == OL_CLASS_CONTACT:
                    contact_count += 1
                elif class_id == OL_CLASS_DISTLIST:
                    distribution_list_count += 1

            folders.append(
                {
                    **folder_meta,
                    "contact_count": contact_count,
                    "distribution_list_count": distribution_list_count,
                    "item_count": contact_count + distribution_list_count,
                }
            )

    folders.sort(key=lambda item: (item["store_name"].lower(), item["folder_path"].lower()))
    return folders


def collect_store_summaries() -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    with outlook_session() as session:
        stores = getattr(session, "Stores", None)
        store_count = int(getattr(stores, "Count", 0) or 0)
        for store_index in range(1, store_count + 1):
            store = stores.Item(store_index)
            store_name = _safe_text(_safe_property(store, "DisplayName")) or f"Store {store_index}"
            try:
                root_folder = store.GetDefaultFolder(OL_FOLDER_CONTACTS)
                contact_root_path = _safe_text(_safe_property(root_folder, "Name")) or "Contacts"
                available = True
            except Exception:
                contact_root_path = ""
                available = False

            summaries.append(
                {
                    "display_name": store_name,
                    "file_path": _safe_text(_safe_property(store, "FilePath")),
                    "exchange_store_type": _safe_text(_safe_property(store, "ExchangeStoreType")),
                    "contact_root_available": available,
                    "contact_root_path": contact_root_path,
                }
            )

    summaries.sort(key=lambda item: item["display_name"].lower())
    return summaries


def search_contact_rows(
    query: str,
    filters: ContactFilters | None = None,
    *,
    limit: int = 25,
) -> list[dict[str, str]]:
    normalized_query = _normalize(query)
    if not normalized_query:
        return []

    results: list[dict[str, str]] = []
    for row in collect_contact_rows(filters):
        haystack = "\n".join(
            [
                row["full_name"],
                row["file_as"],
                row["company_name"],
                row["email_1_address"],
                row["email_2_address"],
                row["email_3_address"],
                row["mobile_phone"],
                row["business_phone"],
                row["home_phone"],
                row["categories"],
                row["notes"],
                row["members"],
            ]
        ).lower()
        if normalized_query in haystack:
            results.append(row)
            if len(results) >= max(1, min(limit, 100)):
                break

    return results


def get_contact_by_identity(
    *,
    entry_id: str | None = None,
    full_name: str | None = None,
    filters: ContactFilters | None = None,
) -> dict[str, Any]:
    if not entry_id and not full_name:
        raise ValueError("Provide entry_id or full_name.")

    matches: list[dict[str, str]] = []
    for row in collect_contact_rows(filters):
        if entry_id and row["entry_id"] == entry_id:
            matches.append(row)
        elif full_name and _normalize(row["full_name"]) == _normalize(full_name):
            matches.append(row)
        elif full_name and _normalize(row["file_as"]) == _normalize(full_name):
            matches.append(row)

    return {
        "match_count": len(matches),
        "matches": matches[:20],
    }


def _rows_for_preview(
    rows: list[dict[str, str]],
    *,
    include_notes_preview: bool,
    preview_chars: int,
) -> list[dict[str, str]]:
    preview_rows: list[dict[str, str]] = []
    for row in rows:
        item = dict(row)
        notes = item.pop("notes", "")
        if include_notes_preview:
            item["notes_preview"] = _truncate(notes, preview_chars)
        preview_rows.append(item)
    return preview_rows


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=CONTACT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contacts"
    sheet.append(CONTACT_COLUMNS)

    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for column_index, header in enumerate(CONTACT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = header_font

    for row in rows:
        sheet.append([row.get(column, "") for column in CONTACT_COLUMNS])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_index, header in enumerate(CONTACT_COLUMNS, start=1):
        column_letter = get_column_letter(column_index)
        max_length = len(header)
        for cell in sheet[column_letter]:
            cell.alignment = wrap_alignment
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    workbook.save(path)


def _write_summary(path: Path, rows: list[dict[str, str]], filters: ContactFilters) -> None:
    contact_count = sum(1 for row in rows if row["item_type"] == "contact")
    distribution_list_count = sum(
        1 for row in rows if row["item_type"] == "distribution_list"
    )
    store_counts = Counter(row["store_name"] for row in rows)
    folder_counts = Counter(f"{row['store_name']} :: {row['folder_path']}" for row in rows)

    lines = [
        "# Outlook Contacts Export Summary",
        "",
        f"- Generated at: `{utc_now_iso()}`",
        f"- Rows exported: `{len(rows)}`",
        f"- Contacts: `{contact_count}`",
        f"- Distribution lists: `{distribution_list_count}`",
        f"- Store filter: `{filters.store_name or 'all'}`",
        f"- Folder filter: `{filters.folder_path or 'all'}`",
        "",
        "## Stores",
        "",
    ]

    if store_counts:
        for store_name, count in sorted(store_counts.items(), key=lambda item: item[0].lower()):
            lines.append(f"- `{store_name}`: `{count}` rows")
    else:
        lines.append("- No rows exported.")

    lines.extend(["", "## Folders", ""])
    if folder_counts:
        for folder_name, count in sorted(folder_counts.items(), key=lambda item: item[0].lower()):
            lines.append(f"- `{folder_name}`: `{count}` rows")
    else:
        lines.append("- No folders exported.")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def export_contacts_snapshot(
    *,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    filters: ContactFilters | None = None,
    clean: bool = False,
) -> dict[str, Any]:
    filters = filters or ContactFilters()

    if clean and output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    rows = collect_contact_rows(filters)
    contact_count = sum(1 for row in rows if row["item_type"] == "contact")
    distribution_list_count = sum(
        1 for row in rows if row["item_type"] == "distribution_list"
    )

    json_path = output_dir / "contacts.json"
    csv_path = output_dir / "contacts.csv"
    xlsx_path = output_dir / "contacts.xlsx"
    manifest_path = output_dir / "manifest.json"
    summary_path = output_dir / "summary.md"

    _write_json(json_path, rows)
    _write_csv(csv_path, rows)
    _write_xlsx(xlsx_path, rows)
    _write_summary(summary_path, rows, filters)

    manifest = {
        "generated_at": utc_now_iso(),
        "platform": platform.platform(),
        "row_count": len(rows),
        "contact_count": contact_count,
        "distribution_list_count": distribution_list_count,
        "store_filter": filters.store_name,
        "folder_filter": filters.folder_path,
        "include_distribution_lists": filters.include_distribution_lists,
        "files": {
            "contacts_json": str(json_path),
            "contacts_csv": str(csv_path),
            "contacts_xlsx": str(xlsx_path),
            "summary_md": str(summary_path),
        },
    }
    _write_json(manifest_path, manifest)
    manifest["files"]["manifest_json"] = str(manifest_path)
    return manifest


def preview_contacts(
    *,
    filters: ContactFilters | None = None,
    limit: int = 25,
    offset: int = 0,
    include_notes_preview: bool = False,
    preview_chars: int = 300,
) -> dict[str, Any]:
    rows = collect_contact_rows(filters)
    page = rows[offset : offset + max(1, min(limit, 100))]
    preview = _rows_for_preview(
        page,
        include_notes_preview=include_notes_preview,
        preview_chars=preview_chars,
    )
    return {
        "count": len(preview),
        "total": len(rows),
        "offset": offset,
        "limit": limit,
        "contacts": preview,
    }
